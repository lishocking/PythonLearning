import sqlite3
import win32com
import os
import openpyxl
import datetime
import pandas as pd
import time
import pdb
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
conn=sqlite3.connect("finance.sql3")
c=conn.cursor()

os.chdir('./report')
workbook=openpyxl.Workbook()

#获得当前时间时间戳 
now = int(time.time()) 
#转换为其他日期格式,如:"%Y-%m-%d %H:%M:%S" 
timeStruct = time.localtime(now) 
strTime = time.strftime("%Y_%m_%d_%H_%M_%S", timeStruct) 

wb = Workbook()
ws1 = wb.active
ws1.title = '总结'
#select max(交割日期),证券名称,剩余数量 from fr  GROUP BY 证券名称
com="select 证券名称,证券代码,sum(清算金额) as 结算总和 from fr group by 证券名称"
c.execute(com)
ws1.append(['证券名称','证券代码','总盈亏','剩余持仓','现在价格','持仓价格','合并盈利'])
count=1
for i in c:
    if i == '':
        continue
    count+=1
    print(i)
    ws1['A%i'%count]=i[0]
    ws1['B%i'%count]=i[1]
    ws1['C%i'%count]='%0.2f'%i[2]
    wb.create_sheet(title=i[0])
for i in range(2,count+1):
    
    name=ws1['A%i'%i].value
    if name == "":continue
    com='select sum(成交数量) from fr  where 证券名称="%s" and 业务名称="证券买入"'%name
    num=c.execute(com)
    buy=num.fetchone()[0]
    if not isinstance(buy,int) : buy=0
    com='select sum(成交数量) from fr  where 证券名称="%s" and 业务名称="证券卖出"'%name
    c.execute(com)
    sell=num.fetchone()[0]
    if not isinstance(sell,int) : sell=0
    #pdb.set_trace()
    last=buy-sell
    ws1['D%i'%i]=last

ws_l=wb.worksheets
for i in ws_l:
    if i.title == '总结':
        continue
    if i.title == 'Sheet':
        continue
    try:
        #pdb.set_trace()
        com="select * from fr where 证券名称 ='%s'"%i.title
        df= pd.read_sql_query(com,conn)
        for r in dataframe_to_rows(df, index=True, header=True):
            i.append(r)
    except:
        pass

conn.close()
wb.save(filename=strTime+'.xlsx')

